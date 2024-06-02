from pathlib import Path
from openpyxl import load_workbook

class Excel_Driver:
    def __init__(self, path: Path) -> None:
        if self.file_exists(path):
            self.path = path
            self.data = []
            self.column_names:list[str] = []
            print(f"Excel Driver created for {self.path}")
        else:
            raise FileExistsError("The file does not exist in Path")

    def __str__(self) -> str:
        return f"Excel_driver for {str(self.path)}"

    def __repr__(self) -> str:
        return f"Excel_Driver({self.path})"

    def _get_data(self):
        # Load the workbook
        workbook = load_workbook(filename=self.path)
        # Select a worksheet
        sheet = workbook.active  # This will select the first sheet as active

        # Iterate through the rows and read the values
        for row in sheet.iter_rows(values_only=True):
            self.data.append(row)

    def file_exists(self, path: Path) -> bool:
        """
        Check if the Excel file exists at the specified path.

        Returns:
        - bool: True if the file exists, False otherwise.
        """
        return path.exists() and path.is_file()
    
    def read(self):
        '''
        Read the data from excel file.
        
        Returns:
        - list: list of data from excel file
        '''
        self.data = []
        self._get_data()
        return self.data
    
    def select(self, selection_strategy = None):
        '''
        Select data from the file using selection_strategy
        selection_strategy is a function which returns a list of data tuples

        Returns:
        - list: Data that have been chosen
        '''
        #IF strategy has been selected use it
        if selection_strategy is not None:
            return selection_strategy(self.data)
        else:
        #ELSE return data unchanged (Identity Function)
            return self.data
    def get_column_names(self):
        return self.column_names
    def set_column_names(self, new_column_names:list[str]):
        self.column_names = new_column_names

if __name__ == '__main__':
    #Strategy for testing data.xlsx
    def strategy(mylist):
        return [x for x in mylist[1:] if x[2]>50210]
    #Strategy for testing fakenames.xlsx (name, surname, True/False)
    def only_true(mylist):
        return [item for item in mylist[1:] if item[-1] ]
    print("START")
    static_dir = Path(__file__).resolve().parent.parent / 'static'
    path = static_dir / 'fakenames.xlsx'
    try:
        driver = Excel_Driver(path)
    except FileExistsError as e:
        print(e)
    driver.read()
    driver.select()
    print(driver.select(only_true))